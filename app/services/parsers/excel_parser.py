"""Excel (XLSX) 解析器。"""

import openpyxl

from app.services.parsers.models import ParsedDocument

_MAX_ROWS = 10000


class ExcelParser:
    @staticmethod
    def parse(file_path: str) -> ParsedDocument:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        full_text_parts = []
        total_rows = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= _MAX_ROWS:
                    sheet_rows.append("...(truncated)")
                    break
                cells = [str(c) if c is not None else "" for c in row]
                sheet_rows.append(" | ".join(cells))
                total_rows += 1
            full_text_parts.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(sheet_rows))

        wb.close()
        return ParsedDocument(
            text="\n\n".join(full_text_parts),
            metadata={"format": "xlsx", "sheets": wb.sheetnames, "row_count": total_rows},
        )


def parse_xlsx(file_path: str) -> ParsedDocument:
    return ExcelParser.parse(file_path)
